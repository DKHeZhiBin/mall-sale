import pandas as pd
from openpyxl import load_workbook
import random

# 读取Excel文件，假设文件名为'漢字.xlsx'
file_path = '漢字.xlsx'
sheetName = 'n2'
df = pd.read_excel(file_path, sheet_name=sheetName, header=None)

# 已测试的行集合
tested_indices = set()
# 错误的行集合
error_indices = set()


def test_word_function_1():
    """功能1：给出单词意思，让用户写出完整单词进行比对"""
    while len(tested_indices) < 10:
        idx = random.randint(0, len(df) - 1)
        if idx not in tested_indices and df.iloc[idx, 12] == 0:
            tested_indices.add(idx)

            meaning = df.iloc[idx, 4]  # 第五列：单词意思
            correct_word = df.iloc[idx, 0]  # 第一列：完整单词
            kanji_reading = df.iloc[idx, 2]  # 第三列：汉字读音
            
            user_input = input(f"请输入意思为 '{meaning}' 的日语单词：")

            if user_input == correct_word:
                print("正确！")
            else:
                error_indices.add(idx)
                print(f"错误，正确答案是: {correct_word}, 读音：{kanji_reading}")

    save_error_indices()


def test_word_function_2():
    """功能2：给出单词意思，让用户写出单词中汉字的读音进行比对"""
    while len(tested_indices) < 50:
        idx = random.randint(0, len(df) - 1)
        if idx not in tested_indices and str(df.iloc[idx, 12]) == "0":
            tested_indices.add(idx)

            meaning = df.iloc[idx, 4]  # 单词意思
            word = df.iloc[idx, 0]  # 单词
            reading = df.iloc[idx, 2]  #读音

            user_input = input(f"请输入意思为 '{meaning}' 的单词：")

            if user_input == word:
                print("正确！")
            else:
                error_indices.add(idx)
                print(f"错误，正确答案是: {word}, 读音是 {reading}")

    save_error_indices2()


def test_word_function_3():
    """功能3：给出单词和意思，让用户写出单词中汉字的读音进行比对"""
    while len(tested_indices) < 50:
        idx = random.randint(0, len(df) - 1)
        if idx not in tested_indices and df.iloc[idx, 10] == 0:
            tested_indices.add(idx)

            word = df.iloc[idx, 0]  # 完整单词
            meaning = df.iloc[idx, 4]  # 单词意思
            kanji_reading = df.iloc[idx, 1]  # 汉字读音

            user_input = input(f"单词 '{word}' 的意思是 '{meaning}'，请输入单词中汉字的读音：")

            if user_input == kanji_reading:
                print("正确！")
            else:
                error_indices.add(idx)
                print(f"错误，正确答案是: {kanji_reading}")

    save_error_indices3()


def test_and_remove_errors_function_4():
    """测试用户是否可以正确回答错误的单词，如果正确，则从工作表中删除该行"""

    workbook = load_workbook(file_path)

    # 检查是否存在'Error'工作表
    if 'Error' not in workbook.sheetnames:
        print("没有找到 'Error' 工作表。")
        return

    sheet = workbook['Error']
    max_row = min(sheet.max_row, 10)

    rows_to_remove = []

    for row in range(1, max_row + 1):
        word = sheet.cell(row=row, column=1).value
        meaning = sheet.cell(row=row, column=4).value
        kanji_reading = sheet.cell(row=row, column=3).value

        if word is None:
            continue

        user_input = input(f"单词 '{word}' 的意思是 '{meaning}'，请输入单词中汉字的读音：")

        if user_input == kanji_reading:
            print("正确！")
            rows_to_remove.append(row)
        else:
            print(f"错误，正确答案是: {kanji_reading}")

    for row in reversed(rows_to_remove):
        print(f"删除行: {row}")
        sheet.delete_rows(row, 1)

    workbook.save(file_path)
    print("已从 'Error' 表中删除答对的单词。")


def save_error_indices2():
    """将错误的单词保存到新的工作表中"""

    book = load_workbook(file_path)

    # 如果没有 Error 表，则创建
    if 'Error' not in book.sheetnames:
        error_sheet = book.create_sheet('Error')
    else:
        error_sheet = book['Error']

    # 写入错误单词
    start_row = error_sheet.max_row + 1
    for i, idx in enumerate(error_indices, start=start_row):
        error_sheet[f'A{i}'] = df.iloc[idx, 0]
        error_sheet[f'B{i}'] = df.iloc[idx, 1]
        error_sheet[f'C{i}'] = df.iloc[idx, 2]
        error_sheet[f'D{i}'] = df.iloc[idx, 4]

    # 将已测试记录写入原工作表
    sheet = book[sheetName]
    for i in tested_indices:
        sheet[f'M{i + 1}'] = 1

    book.save(file_path)
    print("错误与已测记录已保存。")

def save_error_indices3():
    """将错误的单词保存到新的工作表中"""

    book = load_workbook(file_path)

    # 如果没有 Error 表，则创建
    if 'Error' not in book.sheetnames:
        error_sheet = book.create_sheet('Error')
    else:
        error_sheet = book['Error']

    # 写入错误单词
    start_row = error_sheet.max_row + 1
    for i, idx in enumerate(error_indices, start=start_row):
        error_sheet[f'A{i}'] = df.iloc[idx, 0]
        error_sheet[f'B{i}'] = df.iloc[idx, 1]
        error_sheet[f'C{i}'] = df.iloc[idx, 2]
        error_sheet[f'D{i}'] = df.iloc[idx, 4]

    # 将已测试记录写入原工作表
    sheet = book[sheetName]
    for i in tested_indices:
        sheet[f'K{i + 1}'] = 1

    book.save(file_path)
    print("错误与已测记录已保存。")

def main():
    while True:
        print("\n选择一个功能:")
        print("1. 根据单词意思，输入完整单词")
        print("2. 根据单词意思，输入汉字读音")
        print("3. 根据单词和意思，输入汉字读音")
        print("4. 重考错误单词")
        print("5. 退出")
        
        choice = input("请输入你的选择 (1-5): ")

        if choice == '1':
            test_word_function_1()
        elif choice == '2':
            test_word_function_2()
        elif choice == '3':
            test_word_function_3()
        elif choice == '4':
            test_and_remove_errors_function_4()
        elif choice == '5':
            print("退出程序")
            break
        else:
            print("无效选择，请重新输入。")


if __name__ == "__main__":
    main()
